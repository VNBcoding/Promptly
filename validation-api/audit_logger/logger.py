"""
audit_logger/logger.py

PRIVACY & GOVERNANCE NOTICE:
This module stores employee prompt content for IT compliance review.
This constitutes administrative employee monitoring and introduces
significant privacy, data retention, and governance obligations.

Before enabling this feature in production:
  1. Consult your legal and HR teams.
  2. Establish a data retention and deletion policy.
  3. Restrict access to audit log files to authorized IT personnel only.
  4. Notify employees that prompts submitted via AI tools are logged.

To disable this module entirely: set AUDIT_LOG_ENABLED=false in environment
or in config.py. The logging call in app.py checks this flag.

DESIGN: Append-only, one row per processed prompt. Uses openpyxl for .xlsx.
"""

import os
import threading
from datetime import datetime, timezone
from typing import Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Thread lock to prevent concurrent write corruption
_write_lock = threading.Lock()

# Column definitions — order matters for the spreadsheet layout
COLUMNS = [
    "Timestamp",
    "Platform",
    "PromptText",
    "AttachmentText",
    "Status",
    "ViolationTypes",
    "Reasons",
    "AutocorrectApplied",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Row colour coding by status
ROW_FILLS = {
    "valid":                      PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    "certain parts need correction": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # red
}


def _ensure_workbook(filepath: str) -> None:
    """
    Create the workbook and header row if the file does not exist.
    """
    if os.path.exists(filepath):
        return

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prompt Audit Log"

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths for readability
    column_widths = {
        "Timestamp": 22,
        "Platform": 22,
        "PromptText": 60,
        "AttachmentText": 50,
        "Status": 22,
        "ViolationTypes": 35,
        "Reasons": 50,
        "AutocorrectApplied": 20,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = column_widths.get(col_name, 20)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"  # Freeze header row

    wb.save(filepath)


def append_audit_entry(
    filepath: str,
    domain: str,
    prompt_text: str,
    attachment_text: str,
    status: str,
    reasons: Optional[List[str]] = None,
    violation_types: Optional[List[str]] = None,
    autocorrect_applied: bool = False,
) -> None:
    """
    Append a single audit log entry to the Excel workbook.

    Args:
        filepath: Absolute path to the .xlsx audit log file.
        domain: The AI platform domain (e.g., "chatgpt.com").
        prompt_text: The typed prompt text.
        attachment_text: Extracted text from attached files (may be empty).
        status: One of "valid" | "certain parts need correction".
        reasons: List of violation reasons (empty list if valid).
        autocorrect_applied: True if the user accepted autocorrect.
    """
    with _write_lock:
        try:
            _ensure_workbook(filepath)
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
            reasons_str = "; ".join(reasons) if reasons else ""
            violation_types_str = "; ".join(violation_types) if violation_types else ""

            row_data = [
                timestamp,
                domain,
                _truncate_for_log(prompt_text, 5000),
                _truncate_for_log(attachment_text, 5000),
                status,
                violation_types_str,
                reasons_str,
                "Yes" if autocorrect_applied else "No",
            ]

            ws.append(row_data)

            # Apply text wrapping and colour coding to the new row
            new_row = ws.max_row
            row_fill = ROW_FILLS.get(status.lower())
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=new_row, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_fill:
                    cell.fill = row_fill

            wb.save(filepath)

        except Exception as exc:
            # Log to stderr but do not raise — audit failure must not block validation
            import sys
            print(f"[ACG Audit] WARNING: Failed to write audit log entry: {exc}", file=sys.stderr)


def _truncate_for_log(text: str, max_chars: int) -> str:
    """Truncate text for the audit log to avoid massive cells."""
    if not text:
        return ""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + f"\n... [truncated at {max_chars} chars]"
