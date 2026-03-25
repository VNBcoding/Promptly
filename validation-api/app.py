"""
app.py
AI Compliance Guard — Local Validation API
Flask application entry point.

Endpoints:
  POST /validate                    — Validate a prompt against org policy
  POST /audit/autocorrect-accepted  — Update latest audit entry: autocorrect was accepted
  GET  /health                      — Health check

All validation and logging happen locally. No data is sent externally.
"""

import os
import sys
from flask import Flask, request, jsonify
from flask_cors import CORS

from config import Config
from validator.rule_based import RuleBasedValidator
from validator.base import ValidatorBase
from audit_logger.logger import append_audit_entry

app = Flask(__name__)
CORS(app, origins=Config.ALLOWED_ORIGINS)

# ---------------------------------------------------------------------------
# Validator factory
# Swap VALIDATOR_BACKEND in config.py to change the validation strategy.
# ---------------------------------------------------------------------------
def get_validator() -> ValidatorBase:
    backend = Config.VALIDATOR_BACKEND
    if backend == "rule_based":
        return RuleBasedValidator()
    elif backend == "llm":
        from validator.llm_validator import LLMValidator
        return LLMValidator(api_url=Config.LLM_API_URL, model=Config.LLM_MODEL_NAME)
    raise ValueError(f"Unknown validator backend: {backend}")


# Instantiate once at startup (validators are stateless)
_validator = get_validator()

# Audit log file path
_audit_log_path = os.path.join(Config.AUDIT_LOG_DIR, Config.AUDIT_LOG_FILENAME)

# In-memory store for the last logged entry per domain (for autocorrect updates)
# Key: domain string, Value: row index in xlsx (1-based, accounting for header)
_last_audit_row: dict = {}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/health", methods=["GET"])
def health():
    """Simple health check endpoint."""
    return jsonify({"status": "ok", "validator_backend": Config.VALIDATOR_BACKEND})


@app.route("/validate", methods=["POST"])
def validate():
    """
    Validate a user prompt against the organization policy.

    Request body (JSON):
      {
        "prompt_text": "...",
        "attachment_text": "...",
        "attachment_notes": ["..."],
        "policy_text": "...",
        "domain": "chatgpt.com"
      }

    Response body (JSON):
      {
        "status": "valid" | "invalid - autocorrect",
        "message": "...",
        "reasons": ["..."],
        "corrected_prompt": "...",
        "violations": [{ "type": "...", "text": "...", "reason": "..." }]
      }
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Request body must be JSON."}), 400

    prompt_text = data.get("prompt_text", "").strip()
    attachment_text = data.get("attachment_text", "").strip()
    attachment_notes = data.get("attachment_notes", [])
    policy_text = data.get("policy_text", "").strip()
    domain = data.get("domain", "unknown")

    if not policy_text:
        return jsonify({"error": "No policy text provided in request."}), 400

    if not prompt_text and not attachment_text:
        return jsonify({"error": "No prompt or attachment text provided."}), 400

    # Run validation
    try:
        result = _validator.validate(
            prompt_text=prompt_text,
            attachment_text=attachment_text,
            policy_text=policy_text,
            domain=domain,
        )
    except Exception as exc:
        app.logger.error(f"Validation error: {exc}", exc_info=True)
        return jsonify({"error": f"Internal validation error: {str(exc)}"}), 500

    # Map internal status to audit log status label
    audit_status = "valid" if result["status"] == "valid" else "certain parts need correction"

    # Append audit log entry
    if Config.AUDIT_LOG_ENABLED:
        violation_types = [v.get("type", "") for v in result.get("violations", [])]
        append_audit_entry(
            filepath=_audit_log_path,
            domain=domain,
            prompt_text=prompt_text,
            attachment_text=attachment_text,
            status=audit_status,
            reasons=result.get("reasons", []),
            violation_types=violation_types,
            autocorrect_applied=False,  # Will be updated if user accepts
        )

    return jsonify(result), 200


@app.route("/audit/autocorrect-accepted", methods=["POST"])
def autocorrect_accepted():
    """
    Called by the extension when a user accepts autocorrect.
    Updates the most recent audit log entry's AutocorrectApplied column.

    NOTE: In MVP, we re-write the last row for the given domain.
    A production implementation would use a proper database with row IDs.
    """
    if not Config.AUDIT_LOG_ENABLED:
        return jsonify({"status": "audit_disabled"}), 200

    data = request.get_json(silent=True) or {}
    domain = data.get("domain", "unknown")

    try:
        _update_last_autocorrect_entry(_audit_log_path, domain)
    except Exception as exc:
        app.logger.warning(f"Could not update autocorrect flag in audit log: {exc}")

    return jsonify({"status": "ok"}), 200


def _update_last_autocorrect_entry(filepath: str, domain: str) -> None:
    """
    Find the most recent row for this domain in the audit log and
    set AutocorrectApplied to 'Yes'.
    """
    import openpyxl
    if not os.path.exists(filepath):
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Column indices (1-based)
    col_map = {cell.value: cell.column for cell in ws[1]}
    domain_col = col_map.get("Platform")
    autocorrect_col = col_map.get("AutocorrectApplied")

    if not domain_col or not autocorrect_col:
        return

    # Scan from bottom to find the most recent row matching domain
    for row in reversed(list(ws.iter_rows(min_row=2))):
        if row[domain_col - 1].value == domain:
            row[autocorrect_col - 1].value = "Yes"
            break

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"[ACG] Starting validation API on {Config.HOST}:{Config.PORT}")
    print(f"[ACG] Validator backend: {Config.VALIDATOR_BACKEND}")
    print(f"[ACG] Audit logging: {'ENABLED' if Config.AUDIT_LOG_ENABLED else 'DISABLED'}")
    if Config.AUDIT_LOG_ENABLED:
        print(f"[ACG] Audit log path: {_audit_log_path}")
    app.run(host=Config.HOST, port=Config.PORT, debug=Config.DEBUG)
